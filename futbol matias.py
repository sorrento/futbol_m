# ---
# jupyter:
#   jupytext:
#     formats: ipynb,py
#     text_representation:
#       extension: .py
#       format_name: light
#       format_version: '1.5'
#       jupytext_version: 1.13.7
#   kernelspec:
#     display_name: Python 3 (ipykernel)
#     language: python
#     name: python3
# ---

# Primero actualizar el excel que lee de las páginas web

# +
import pandas as pd
import numpy as np
from openpyxl import load_workbook

path = 'data_in/Resultados futbol matías.xlsx'
wb = load_workbook(filename=path)
s = wb.sheetnames
# -

acc = pd.DataFrame()
for name in s:
    try:
        df = pd.read_excel(path, sheet_name=name)
        fecha = df.columns[-1][:-1]
        jornada = df.columns[0].split(' ')[1]
        df = df.iloc[:, [0, 2, 4, 6]].copy()
        df.columns = ['Local', 'goles_l', 'goles_v', 'Visita']
        df['fecha'] = fecha
        df['jornada'] = int(jornada)
        acc = pd.concat([acc, df])
    except:
        print('fail:', name)

acc = acc.dropna().copy()
acc['dif'] = acc.goles_l - acc.goles_v
acc

equipos = sorted(list(set(list(acc.Local) + list(acc.Visita))))
equipos

# # Creación de la matriz de diferencias

# +
names = ['Local', 'Visita', 'dif', 'fecha', 'jornada']

acc2 = acc.copy()  # invertimos local y visita
acc2 = acc2.rename(columns={'Local': 'Visita', 'Visita': 'Local'})
acc2['dif'] = -acc2.dif

acct = pd.concat([acc, acc2])[names]
acct['vic'] = np.where(acct.dif > 0, 1, 0)
acct['ptos'] = np.where(acct.dif > 0, 3, np.where(acct.dif < 0, 0, 1))
acct.reset_index(drop=True)
acct
# -

jornadas = acct.jornada.drop_duplicates()

# +
# ranking usando la información hasta la jornada j
ran_acc = pd.DataFrame()
di = {}

for j in jornadas:
    acu = acct[acct.jornada <= j]
    ran = acu.groupby('Local').agg({'dif': 'mean', 'vic': 'mean', 'ptos': 'sum'}).sort_values(['ptos', 'dif'],
                                                                                              ascending=False)
    ran['r'] = range(1, len(ran) + 1)
    ran['name_ran'] = list(ran.reset_index().apply(lambda row: str(row.r).zfill(2) + '_' + row.Local, axis=1))
    ran['jornada'] = j

    # sesgo: calculemos la media del ranking (puntos) de los rivales de cada equipo
    ptos_rival = acu[['Local', 'Visita']].merge(
        ran.reset_index()[['Local', 'ptos']].rename(columns={'Local': 'Visita'}),
        on='Visita')
    ptos_rival = ptos_rival.groupby('Local').mean().sort_values('ptos', ascending=False).rename(
        columns={'ptos': 'ptos_rival'})
    ran = ran.join(ptos_rival)

    dic_names = ran[['name_ran']].T.to_dict('records')[0]

    di[j] = dic_names
    ran_acc = pd.concat([ran_acc, ran])
# -

ran_acc

buf = pd.DataFrame().from_dict(di, orient='index').reset_index().melt(id_vars='index', value_name='name_ra',
                                                                      var_name='equipo').drop_duplicates().rename(
    columns={'index': 'jornada'})
buf

def popo(acct, buf, x):
    return acct.merge(buf, left_on=['jornada', x], right_on=['jornada', 'equipo']).drop(columns='equipo').rename(
        columns={'name_ra': f'{x}_ran'})


acct = popo(acct, buf, 'Local')
acct = popo(acct, buf, 'Visita')
acct = acct.sort_values(['jornada', 'Local_ra'], ascending=[False, True])

acct

acct.to_excel('data.xlsx', index=False)
ran.to_excel('data_ranking.xlsx', index=True)

# partidos de un equipo
acc[acc.Visita == 'JUNIOR,C.F. F']

# ## otro

loc = acc.groupby('Local').mean().rename(columns={'goles_l': 'g_f_L', 'goles_v': 'g_c_L'})
vis = acc.groupby('Visita').mean().rename(columns={'goles_v': 'g_f_V', 'goles_l': 'g_c_V'})

goles = loc.join(vis)
goles

goles.to_excel('goles.xlsx')

# +
# usar matriz de heatmap con dif goles por partido

# +
# efecto local vs visita
# -

# las estadisticas pueden estar sesgadas por el órden de los partidos "te pueden habe tocado primero los buenos"

# +
# Minimización de reanking universal dado por diferencia de golesa
# No# consideramos local o visita

# +
acc['dif'] = acc.goles_l - acc.goles_v

pares = acc[['Local', 'Visita', 'dif']].copy()
pares  # positivo es que gana el local

# varis=['x_'+str(i) for i in range(len(equipos))]
varis = [i for i in range(len(equipos))]

di = dict(zip(equipos, varis))

pares['xl'] = pares.apply(lambda x: di[x.Local], axis=1)
pares['xv'] = pares.apply(lambda x: di[x.Visita], axis=1)

pares

# +
# construir la funcion a miminizar
# -

# # Completando la matriz:

pares

# +
import numpy as np

n = len(equipos)

arr = np.full((n, n), np.nan)
arr2 = np.full((n, n), np.nan)

for _, row in pares.iterrows():
    arr[row.xl, row.xv] = row.dif

    gana = np.where(row.dif > 0, 1, 0)
    arr2[row.xl, row.xv] = gana
# -

pd.DataFrame(arr)

pd.DataFrame(arr2)

# +
m_fila = np.nanmean(arr, axis=1)
m_col = np.nanmean(arr, axis=0)

m_fila2 = np.nanmean(arr2, axis=1)
m_col2 = np.nanmean(arr2, axis=0)
# -

for i in range(n):
    for j in range(n):
        if np.isnan(arr[i, j]):
            arr[i, j] = (m_fila[i] + m_col[j]) / 2
            # probs de victoria
            arr2[i, j] = (m_fila2[i] + m_col2[j]) / 2

pd.DataFrame(arr2)


def get_name(pat):
    li = [x for x in equipos if pat.upper() in x]
    if len(li) == 0:
        print('error')
        res = ''
    else:
        res = li[0]
    return res


def predice(eq1, eq2):
    e1 = get_name(eq1)
    e2 = get_name(eq2)

    i, j = di[e1], di[e2]

    print(f'***** {e1} vs {e2}\n')

    d = round(arr[i, j], 1)
    prob = round(arr2[i, j], 2)
    win = np.where(d > 0, e1, e2)
    print(f'{e1} vs {e2} | gana: {win} por {abs(d)} | prob win local: {prob}')

    d = round(arr[j, i], 1)
    prob = round(arr2[j, i], 2)
    win = np.where(d > 0, e2, e1)
    print(f'{e2} vs {e1} | gana: {win} por {abs(d)} | prob win local: {prob}')


predice('cugat', 'egara')

predice('cugat', 'juan')

predice('cugat', 'farga')

predice('cugat', 'pueblo')

predice('cugat', 'cavall')

predice('cugat', 'aran')

predice('cugat', 'septi')

predice('cugat', 'junio')

predice('cugat', 'europa')  # j10

predice('cugat', 'oli')  # j11

predice('cugat', 'boada')  # j12

# +
# puede haber sesgo porque los de local han tocado más difíciles
